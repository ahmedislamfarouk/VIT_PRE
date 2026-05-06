import pandas as pd
import openpyxl

# Create workbook
wb = openpyxl.Workbook()

# ============================================================================
# Sheet 1: IID Ranking
# ============================================================================
ws1 = wb.active
ws1.title = "IID Ranking"

headers = ["Rank", "Algorithm", "Objective", "TSR", "Jain", "Energy/Task", "Step Reward", "Throughput", "Fairness-CV", "Entropy", "TSR Stability"]
data = [
    [1, "HiGAT-HAPPO (Full)", 0.741993, "96.88%", 0.997150, 40.286, 0.501416, 493.342, 0.041439, 4.093867, 0.015769],
    [2, "MAPPO (No FL)", 0.688835, "95.69%", 0.993798, 73.072, 0.545478, 488.500, 0.061217, 5.947139, 0.009893],
    [3, "MAT (No FL)", 0.647134, "100.00%", 0.987500, 517.257, 0.020180, 85.260, 0.040799, "N/A", 0.000000],
    [4, "GAT-NoEdge-HAPPO", 0.599178, "95.10%", 0.995524, 48.606, 0.461946, 478.762, 0.058716, 4.070937, 0.015825],
    [5, "FlatGAT-HAPPO", 0.585549, "95.09%", 0.995808, 54.675, 0.423285, 456.896, 0.054055, 4.058082, 0.017793],
    [6, "HAPPO (No FL)", 0.336867, "100.00%", 0.953646, 909.257, 0.035031, 50.029, 0.136871, "N/A", 0.000000],
]

ws1.append(headers)
for row in data:
    ws1.append(row)

# ============================================================================
# Sheet 2: NONIID Ranking
# ============================================================================
ws2 = wb.create_sheet("NONIID Ranking")

data_noniid = [
    [1, "NI: F2 Only (Energy)", 0.883450, "98.05%", 0.998669, 34.287, 0.504976, 524.612, 0.028299, 7.238111, 0.020273],
    [2, "NI: F2+F3 (Ene+Fair)", 0.878355, "97.68%", 0.998346, 32.869, 0.512848, 546.628, 0.035472, 7.218045, 0.009179],
    [3, "NI: F1+F3 (Rew+Fair)", 0.845637, "96.80%", 0.997664, 29.547, 0.544536, 558.168, 0.042380, 7.225349, 0.011904],
    [4, "MAPPO (No FL)", 0.766562, "96.39%", 0.991269, 63.360, 0.560173, 584.775, 0.071769, 5.420131, 0.006206],
    [5, "NI: F1+F2+F3 (All)", 0.759913, "95.22%", 0.996464, 31.151, 0.513865, 546.038, 0.054153, 7.211203, 0.019471],
    [6, "HiGAT-HAPPO (Full)", 0.703827, "94.00%", 0.995341, 29.536, 0.523093, 552.138, 0.063334, 7.201962, 0.023986],
    [7, "NI: F3 Only (Fairness)", 0.694224, "93.83%", 0.995198, 32.435, 0.495525, 534.809, 0.061826, 7.234883, 0.016683],
    [8, "NI: F1+F2 (Rew+Ene)", 0.672213, "93.47%", 0.994766, 31.599, 0.498968, 539.434, 0.066958, 7.212312, 0.024281],
    [9, "GAT-NoEdge-HAPPO", 0.645412, "92.78%", 0.994400, 29.662, 0.514314, 552.054, 0.067840, 7.169845, 0.027154],
    [10, "FlatGAT-HAPPO", 0.636841, "92.59%", 0.993687, 32.739, 0.489741, 553.623, 0.069447, 7.201206, 0.019129],
    [11, "NI: F1 Only (Reward)", 0.631079, "93.01%", 0.989093, 35.559, 0.468886, 504.155, 0.075542, 7.223282, 0.012348],
    [12, "MAT (No FL)", 0.351102, "100.00%", 0.925260, 681.654, -0.075905, 61.320, 0.218871, "N/A", 0.000000],
    [13, "HAPPO (No FL)", 0.348907, "100.00%", 0.923958, 741.307, 0.003344, 57.672, 0.219714, "N/A", 0.000000],
]

ws2.append(headers)
for row in data_noniid:
    ws2.append(row)

# ============================================================================
# Sheet 3: Cross-Condition Delta
# ============================================================================
ws3 = wb.create_sheet("Cross-Condition Delta")

headers_delta = ["Algorithm", "IID Objective", "NONIID Objective", "Delta Objective", "IID Rank", "NONIID Rank", "Delta TSR (pp)", "Delta Jain", "Delta Energy/Task"]
data_delta = [
    ["FlatGAT", 0.585549, 0.636841, 0.051292, 5, 10, -2.499, -0.002121, -21.936],
    ["GAT-NoEdge", 0.599178, 0.645412, 0.046234, 4, 9, -2.326, -0.001124, -18.944],
    ["HAPPO", 0.336867, 0.348907, 0.012040, 6, 13, 0.000, -0.029688, -167.950],
    ["HiGAT-Full", 0.741993, 0.703827, -0.038166, 1, 6, -2.879, -0.001809, -10.751],
    ["MAPPO", 0.688835, 0.766562, 0.077727, 2, 4, 0.699, -0.002529, -9.712],
    ["MAT", 0.647134, 0.351102, -0.296032, 3, 12, 0.000, -0.062240, 164.397],
]

ws3.append(headers_delta)
for row in data_delta:
    ws3.append(row)

# ============================================================================
# Sheet 4: Final Objective Highlights (IID + NONIID combined)
# ============================================================================
ws4 = wb.create_sheet("Final Obj Highlights")

headers_obj = ["Condition", "Rank", "Algorithm", "Objective", "TSR", "Jain", "Energy/Task", "Step Reward", "Throughput"]
data_obj = [
    ["IID", 1, "HiGAT-HAPPO (Full)", 0.741993, "96.88%", 0.997150, 40.286, 0.501416, 493.342],
    ["IID", 2, "MAPPO (No FL)", 0.688835, "95.69%", 0.993798, 73.072, 0.545478, 488.500],
    ["IID", 3, "MAT (No FL)", 0.647134, "100.00%", 0.987500, 517.257, 0.020180, 85.260],
    ["NONIID", 1, "NI: F2 Only (Energy)", 0.883450, "98.05%", 0.998669, 34.287, 0.504976, 524.612],
    ["NONIID", 2, "NI: F2+F3 (Ene+Fair)", 0.878355, "97.68%", 0.998346, 32.869, 0.512848, 546.628],
    ["NONIID", 3, "NI: F1+F3 (Rew+Fair)", 0.845637, "96.80%", 0.997664, 29.547, 0.544536, 558.168],
    ["NONIID", 4, "MAPPO (No FL)", 0.766562, "96.39%", 0.991269, 63.360, 0.560173, 584.775],
    ["NONIID", 5, "NI: F1+F2+F3 (All)", 0.759913, "95.22%", 0.996464, 31.151, 0.513865, 546.038],
]

ws4.append(headers_obj)
for row in data_obj:
    ws4.append(row)

# ============================================================================
# Sheet 5: Isaac Transfer Summary (Zero/Few/Post-Finetune)
# ============================================================================
ws5 = wb.create_sheet("Isaac Transfer Summary")

headers_isaac = ["Algorithm", "Seeds", "ZS TSR", "ZS Jain", "FT TSR", "PFT TSR", "ΔPFT", "PFT Reward"]
data_isaac = [
    ["MAPPO", 3, "92.1%", 0.6292, "91.8%", "92.5%", "+0.4pp", 53.33],
    ["HiGAT NI-F1F2F3", 3, "76.1%", 0.6510, "78.1%", "78.5%", "+2.4pp", 40.89],
    ["HiGAT no_fl", 3, "75.7%", 0.6448, "77.2%", "77.2%", "+1.5pp", 39.47],
]

ws5.append(headers_isaac)
for row in data_isaac:
    ws5.append(row)

# ============================================================================
# Sheet 6: Isaac Per-Seed Detail
# ============================================================================
ws6 = wb.create_sheet("Isaac Per-Seed Detail")

headers_perseed = ["Algorithm", "Seed", "ZS TSR", "PFT TSR", "ΔPFT", "ZS Jain"]
data_perseed = [
    ["MAPPO", 42, "94.0%", "95.2%", "+1.2pp", 0.679],
    ["MAPPO", 123, "88.8%", "90.4%", "+1.6pp", 0.623],
    ["MAPPO", 456, "93.5%", "91.8%", "-1.8pp", 0.585],
    ["HiGAT NI-F1F2F3", 42, "75.3%", "77.8%", "+2.5pp", 0.696],
    ["HiGAT NI-F1F2F3", 123, "72.5%", "77.3%", "+4.8pp", 0.667],
    ["HiGAT NI-F1F2F3", 456, "80.5%", "80.5%", "-0.0pp", 0.590],
    ["HiGAT no_fl", 42, "72.5%", "76.4%", "+3.9pp", 0.696],
    ["HiGAT no_fl", 123, "72.8%", "75.3%", "+2.6pp", 0.635],
    ["HiGAT no_fl", 456, "81.8%", "79.8%", "-2.0pp", 0.603],
]

ws6.append(headers_perseed)
for row in data_perseed:
    ws6.append(row)

# ============================================================================
# Sheet 7: Isaac Full-Study Evaluation (Corrected)
# ============================================================================
ws7 = wb.create_sheet("Isaac Full-Study")

headers_full = ["Algorithm", "Seeds", "TSR mean", "TSR std", "CV", "Jain", "Energy/ep", "Ene/Task ratio"]
data_full = [
    ["MAPPO", 3, "94.7%", "±2.2%", 0.023, 0.633, "5,174J", 5461],
    ["HiGAT NI-F1F2F3", 3, "76.7%", "±4.4%", 0.058, 0.600, "5,092J", 6637],
    ["HiGAT no_fl", 3, "59.4%", "±10.3%", 0.174, 0.621, "5,042J", 8485],
]

ws7.append(headers_full)
for row in data_full:
    ws7.append(row)

# ============================================================================
# Sheet 8: Isaac Full-Study Per-Seed (Corrected)
# ============================================================================
ws8 = wb.create_sheet("Isaac Per-Seed Full")

headers_perseed_full = ["Algorithm", "Seed", "TSR", "Jain", "Energy/ep", "Ene/Task ratio"]
data_perseed_full = [
    ["MAPPO", 42, "97.7%", 0.672, "5,186J", 5309],
    ["MAPPO", 123, "92.4%", 0.610, "5,167J", 5594],
    ["MAPPO", 456, "94.2%", 0.617, "5,171J", 5489],
    ["HiGAT NI-F1F2F3", 42, "82.0%", 0.675, "5,106J", 6230],
    ["HiGAT NI-F1F2F3", 123, "77.1%", 0.636, "5,086J", 6598],
    ["HiGAT NI-F1F2F3", 456, "71.1%", 0.489, "5,084J", 7148],
    ["HiGAT no_fl", 42, "69.6%", 0.679, "5,060J", 7276],
    ["HiGAT no_fl", 123, "63.5%", 0.549, "5,048J", 7955],
    ["HiGAT no_fl", 456, "45.3%", 0.636, "5,017J", 11085],
]

ws8.append(headers_perseed_full)
for row in data_perseed_full:
    ws8.append(row)

# ============================================================================
# Sheet 9: Isaac HAPPO Finetuning (50 rounds)
# ============================================================================
ws9 = wb.create_sheet("Isaac HAPPO Finetune")

headers_ft = ["Metric", "Value"]
data_ft = [
    ["Rounds Complete", "50/50"],
    ["Best TSR", "77.8% (round 2)"],
    ["Final TSR", "72.6% (round 50)"],
    ["ZS Baseline (seed42)", "82.0%"],
    ["Outcome", "Finetuning HURTS — 72.6% < 82.0%"],
    ["Trajectory", "Peak at round 2, monotonic decline to round 50"],
    ["Root Cause 1", "Random critic at start → noisy GAE advantages"],
    ["Root Cause 2", "Small batch (1000 steps/update) with random resets"],
    ["Root Cause 3", "HAPPO IS weight compounding across sequential updates"],
]

ws9.append(headers_ft)
for row in data_ft:
    ws9.append(row)

# ============================================================================
# Sheet 10: Isaac Objective Comparison - Overall Ranking
# ============================================================================
ws10 = wb.create_sheet("Isaac Obj Ranking")

headers_obj_isaac = ["Rank", "Algorithm", "Isaac Obj", "TSR", "Jain", "Reward", "Rew/kJ", "CompE/task", "Seed Stab"]
data_obj_isaac = [
    [1, "MAPPO", 0.960, "94.7%", 0.633, 56.35, 10.89, "179J", 0.977],
    [2, "HiGAT NI-F1F2F3", 0.346, "76.7%", 0.600, 38.65, 7.59, "113J", 0.942],
    [3, "HiGAT no_fl", 0.202, "59.4%", 0.621, 22.12, 4.39, "63J", 0.826],
]

ws10.append(headers_obj_isaac)
for row in data_obj_isaac:
    ws10.append(row)

# ============================================================================
# Sheet 11: Isaac Component Breakdown
# ============================================================================
ws11 = wb.create_sheet("Isaac Component Breakdown")

headers_comp = ["Algorithm", "TSR (x0.30)", "Fairness (x0.20)", "EnergyEff (x0.15)", "StepRew (x0.12)", "Throughput (x0.10)", "FairStab (x0.07)", "CompEff (x0.04)", "Stability (x0.02)", "Total"]
data_comp = [
    ["MAPPO", 0.300, 0.200, 0.150, 0.120, 0.100, 0.070, 0.000, 0.020, 0.960],
    ["HiGAT NI-F1F2F3", 0.147, 0.000, 0.074, 0.058, 0.040, 0.000, 0.013, 0.015, 0.346],
    ["HiGAT no_fl", 0.000, 0.128, 0.000, 0.000, 0.000, 0.034, 0.040, 0.000, 0.202],
]

ws11.append(headers_comp)
for row in data_comp:
    ws11.append(row)

# ============================================================================
# Sheet 12: Isaac Per-Seed Metric Table
# ============================================================================
ws12 = wb.create_sheet("Isaac Per-Seed Metrics")

headers_ps = ["Algorithm", "Seed", "Isaac Obj", "TSR", "Jain", "Reward", "Rew/kJ", "CompE/ep", "CompE/task", "TotE/ep"]
data_ps = [
    ["MAPPO", 42, "—", "97.7%", 0.672, 60.23, 11.62, "181J", "185J", "5,186J"],
    ["MAPPO", 123, "—", "92.4%", 0.610, 52.76, 10.21, "162J", "176J", "5,167J"],
    ["MAPPO", 456, "—", "94.2%", 0.617, 56.07, 10.84, "166J", "176J", "5,171J"],
    ["HiGAT NI-F1F2F3", 42, "—", "82.0%", 0.675, 44.73, 8.76, "101J", "123J", "5,106J"],
    ["HiGAT NI-F1F2F3", 123, "—", "77.1%", 0.636, 38.45, 7.56, "81J", "105J", "5,085J"],
    ["HiGAT NI-F1F2F3", 456, "—", "71.1%", 0.489, 32.76, 6.44, "79J", "111J", "5,084J"],
    ["HiGAT no_fl", 42, "—", "69.6%", 0.679, 32.40, 6.40, "56J", "80J", "5,060J"],
    ["HiGAT no_fl", 123, "—", "63.5%", 0.549, 23.54, 4.66, "43J", "68J", "5,048J"],
    ["HiGAT no_fl", 456, "—", "45.3%", 0.636, 10.42, 2.08, "13J", "28J", "5,017J"],
]

ws12.append(headers_ps)
for row in data_ps:
    ws12.append(row)

# ============================================================================
# Sheet 13: Metric Flip Training vs Isaac
# ============================================================================
ws13 = wb.create_sheet("Training vs Isaac Flip")

headers_flip = ["Metric", "Training env (NONIID sim)", "Isaac (rigid-body sim)", "Direction"]
data_flip = [
    ["TSR", "HiGAT 95.2% vs MAPPO 96.4%", "HiGAT 76.7% vs MAPPO 94.7%", "Flipped (−18pp gap)"],
    ["Jain fairness", "HiGAT 0.996 vs MAPPO 0.991", "HiGAT 0.600 vs MAPPO 0.633", "Flipped (HiGAT loses)"],
    ["Energy/task", "HiGAT 31J vs MAPPO 63J", "HiGAT 113J vs MAPPO 179J", "Maintained (HiGAT still −37%)"],
    ["CompE/task", "HiGAT wins", "HiGAT still wins (113J vs 179J)", "Maintained"],
    ["Obj score", "MAPPO 0.767 vs HiGAT 0.704", "MAPPO 0.960 vs HiGAT 0.346", "Gap widened"],
]

ws13.append(headers_flip)
for row in data_flip:
    ws13.append(row)

# ============================================================================
# Sheet 14: Projected Performance with Obs Normalization
# ============================================================================
ws14 = wb.create_sheet("Projected w Obs Norm")

headers_proj = ["Metric", "HiGAT NI actual", "HiGAT NI projected (+obs norm)", "MAPPO", "Remaining gap"]
data_proj = [
    ["TSR", "76.7%", "~85.4%", "94.7%", "~9pp"],
    ["Jain", 0.600, "~0.635", 0.633, "≈0"],
    ["CompE/task", "113J", "~97J", "179J", "HiGAT still −46%"],
    ["Reward", 38.65, "~43.0", 56.35, "~13 pts"],
    ["Rew/kJ", 7.59, "~8.45", 10.89, "~2.44"],
    ["Isaac Obj", 0.346, "~0.58 (est.)", 0.960, "~0.38"],
]

ws14.append(headers_proj)
for row in data_proj:
    ws14.append(row)

# ============================================================================
# Sheet 15: Episode TSR Trajectory (Collapse Characterization)
# ============================================================================
ws15 = wb.create_sheet("Episode TSR Trajectory")

headers_traj = ["Algorithm", "Seed", "ep1", "ep2-5 avg", "ep6-30 avg", "ep30", "ep1→ss drop"]
data_traj = [
    ["MAPPO", 42, "100.0%", "98.2%", "97.5%", "98.0%", "−2.4pp"],
    ["MAPPO", 123, "94.4%", "92.2%", "92.3%", "92.6%", "−2.2pp"],
    ["MAPPO", 456, "94.4%", "94.3%", "94.2%", "93.5%", "−0.2pp"],
    ["HiGAT NI-F1F2F3", 42, "90.8%", "84.5%", "81.2%", "81.6%", "−9.1pp"],
    ["HiGAT NI-F1F2F3", 123, "79.2%", "77.0%", "77.0%", "77.0%", "−2.2pp"],
    ["HiGAT NI-F1F2F3", 456, "69.0%", "70.5%", "71.3%", "71.0%", "+2.2pp"],
    ["HiGAT no_fl", 42, "80.0%", "69.9%", "69.1%", "68.4%", "−10.8pp"],
    ["HiGAT no_fl", 456, "36.6%", "41.2%", "46.3%", "46.2%", "+8.9pp"],
]

ws15.append(headers_traj)
for row in data_traj:
    ws15.append(row)

# ============================================================================
# Sheet 16: Variance Decomposition
# ============================================================================
ws16 = wb.create_sheet("Variance Decomposition")

headers_var = ["Algorithm", "Within-seed std", "Across-seed std", "Seed max−min"]
data_var = [
    ["MAPPO", "0.70pp", "2.21pp", "5.3pp"],
    ["HiGAT NI-F1F2F3", "1.37pp", "4.43pp", "10.8pp"],
    ["HiGAT no_fl", "1.84pp", "10.32pp", "24.3pp"],
]

ws16.append(headers_var)
for row in data_var:
    ws16.append(row)

# ============================================================================
# Sheet 17: Jain Fairness Trajectory
# ============================================================================
ws17 = wb.create_sheet("Jain Trajectory")

headers_jain = ["Algorithm", "Seed", "ep1 Jain", "ep1-10", "ep11-30", "std", "trend"]
data_jain = [
    ["MAPPO", 42, 0.5993, 0.6698, 0.6734, 0.018, "+0.004"],
    ["MAPPO", 123, 0.6052, 0.5907, 0.6189, 0.017, "+0.028"],
    ["MAPPO", 456, 0.5763, 0.5885, 0.6317, 0.027, "+0.043"],
    ["HiGAT NI-F1F2F3", 42, 0.6012, 0.6728, 0.6764, 0.018, "+0.004"],
    ["HiGAT NI-F1F2F3", 123, 0.5983, 0.6211, 0.6440, 0.016, "+0.023"],
    ["HiGAT NI-F1F2F3", 456, 0.4455, 0.4588, 0.5038, 0.029, "+0.045"],
]

ws17.append(headers_jain)
for row in data_jain:
    ws17.append(row)

# ============================================================================
# Sheet 18: Energy Breakdown (delta-corrected)
# ============================================================================
ws18 = wb.create_sheet("Energy Breakdown")

headers_energy = ["Algorithm", "Total/ep", "Flight", "Computation", "Communication", "Comp/Task"]
data_energy = [
    ["MAPPO", "5,174J", "5,005J (96.7%)", "169.8J (3.3%)", "0.08J", "179.2J"],
    ["HiGAT NI-F1F2F3", "5,092J", "5,005J (98.3%)", "86.9J (1.7%)", "0.16J", "113.2J"],
    ["HiGAT no_fl", "5,042J", "5,005J (99.3%)", "37.2J (0.7%)", "0.13J", "62.5J"],
]

ws18.append(headers_energy)
for row in data_energy:
    ws18.append(row)

# ============================================================================
# Sheet 19: Episode Duration Anomaly
# ============================================================================
ws19 = wb.create_sheet("Episode Duration")

headers_dur = ["Algorithm", "Duration/ep", "Std", "r(TSR, dur)"]
data_dur = [
    ["MAPPO", "32.31s", "±18.85s", "-0.060"],
    ["HiGAT NI-F1F2F3", "10.64s", "±1.87s", "-0.813"],
    ["HiGAT no_fl", "68.36s", "±18.56s", "+0.529"],
]

ws19.append(headers_dur)
for row in data_dur:
    ws19.append(row)

# ============================================================================
# Sheet 20: Observation Distribution Shift
# ============================================================================
ws20 = wb.create_sheet("Obs Distribution Shift")

headers_obs = ["Category", "Count", "Fraction", "Severity"]
data_obs = [
    ["Nearly constant (std < 0.015)", "34 / 73", "47%", "Amplified 67× by HiGAT input projection"],
    ["Large DC offset (|mean| > 0.5)", "20 / 73", "27%", "Far from training distribution center"],
    ["Wide spread (std > 0.25)", "20 / 73", "27%", "Larger-than-expected variance in Isaac"],
    ["Amplified > 20× (std < 0.05)", "38 / 73", "52%", "Major distortion in GNN node feature computation"],
]

ws20.append(headers_obs)
for row in data_obs:
    ws20.append(row)

# ============================================================================
# Sheet 21: TSR-Jain Correlation
# ============================================================================
ws21 = wb.create_sheet("TSR-Jain Correlation")

headers_corr = ["Algorithm", "r(TSR, Jain)", "Jain range"]
data_corr = [
    ["MAPPO", 0.704, "[0.524, 0.710]"],
    ["HiGAT NI-F1F2F3", 0.874, "[0.400, 0.713]"],
    ["HiGAT no_fl", 0.046, "[0.482, 0.728]"],
]

ws21.append(headers_corr)
for row in data_corr:
    ws21.append(row)

# ============================================================================
# Sheet 22: Root-Cause Summary
# ============================================================================
ws22 = wb.create_sheet("Root-Cause Summary")

headers_rc = ["Root Cause", "Estimated Gap", "Fixable?", "How"]
data_rc = [
    ["Obs distribution shift (47% dead dims)", "~9pp", "Yes", "ObsNormalizer (already coded)"],
    ["Spawn-topology sensitivity (attention coupling)", "~7pp", "Partially", "More fewshot adaptation, domain randomization"],
    ["Seed-level baseline variance (spawn positions)", "~4pp", "Partially", "More seeds (n>3), curriculum spawn diversity"],
    ["Random critic in finetune (prevents convergence)", "~3pp potential gain", "Yes", "Critic pre-warming (10-round freeze)"],
    ["Architecture: attention sensitivity vs MAPPO MLP", "residual ~2pp", "Hard", "Architecture change needed"],
    ["Total attributable gap (HiGAT NI vs MAPPO)", "~18pp", "Mostly", "Fixes 1+4 alone could close ~12pp"],
]

ws22.append(headers_rc)
for row in data_rc:
    ws22.append(row)

# ============================================================================
# Sheet 23: Isaac Few-Shot Transfer Averages
# ============================================================================
ws23 = wb.create_sheet("Isaac Few-Shot Avg")

headers_fs = ["Algorithm", "Seeds", "Zero-shot TSR", "Post-finetune TSR", "TSR Gain", "Zero-shot Reward", "Post-finetune Reward", "Reward Gain", "Zero-shot Jain", "Post-finetune Jain"]
data_fs = [
    ["higat_happo", 3, 0.759827, 0.777919, 0.018092, 39.134040, 40.406984, 1.272944, 0.661420, 0.670606],
    ["mappo", 3, 0.921160, 0.924778, 0.003618, 54.183767, 53.334696, -0.849071, 0.629220, 0.646349],
]

ws23.append(headers_fs)
for row in data_fs:
    ws23.append(row)

# ============================================================================
# Sheet 24: Q1-Readiness Assessment
# ============================================================================
ws24 = wb.create_sheet("Q1 Readiness")

headers_q1 = ["Category", "Status", "Notes"]
data_q1 = [
    ["Multi-condition evidence (IID+NONIID)", "Strong", "Clear metric breadth and ablation depth"],
    ["Isaac transfer experiments", "Complete", "3 algos x 3 seeds x 3 phases + full-study"],
    ["NI-fix ablation", "Quantified", "+17.3pp TSR, 3x CV reduction, 24% energy/task reduction"],
    ["HAPPO finetuning negative result", "Clean", "72.6% vs 82.0% ZS, root cause identified"],
    ["Statistical significance testing", "Missing", "No paired tests, CIs, effect sizes"],
    ["Baseline parity", "Issues", "MAT/HAPPO partial metric logging comparability"],
    ["Robustness/sensitivity analysis", "Missing", "No weight sensitivity, seed expansion, HP stability"],
    ["Transfer benchmark breadth", "Limited", "Single Isaac task family, no broader domain shift"],
    ["Computational budget analysis", "Missing", "No normalized training cost/wall-time in claims"],
    ["Isaac gap explanation", "Incomplete", "Obs shift hypothesis not isolated with ablation"],
    ["Overall Q1 readiness", "Not yet", "Close to strong workshop or mid-tier if tightened"],
]

ws24.append(headers_q1)
for row in data_q1:
    ws24.append(row)

# Save
output_path = "/home/ahmed/Desktop/COMPLETE_PROJECT_TABLES.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
